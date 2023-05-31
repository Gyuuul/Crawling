import openpyxl

path=r"C:\pythonStart\파이썬_엑셀다루기\인적사항_data.xlsx"

# 1.엑셀 불러오기
wb= openpyxl.load_workbook(path)

# 2. 엑셀 워크시트 선택
ws = wb['사람']

# 3. 데이터 수정하기
ws['A3'] = "홍길동"
ws['B3'] = "남자"

# 4. 수정 후 엑셀 저장
wb.save(path)
