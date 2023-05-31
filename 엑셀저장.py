import openpyxl

# 1. 엑셀파일 만들기
wb= openpyxl.Workbook()

# 2. 엑셀 워크시트 만들기
ws= wb.create_sheet ('사람')

# 3. 워크시트에 데이터 추가
ws['A1'] = '이름'
ws['B1'] = '성별'

ws['A2'] = '심청'
ws['B2'] = '여자'

# 4. 엑셀에 데이터를 추가했다면 저장!
wb.save(r'C:\pythonStart\파이썬_엑셀다루기\인적사항_data.xlsx')
